"""
Survey routes: list, create, view, add point, export (GeoJSON/XLSX/DXF/GPKG/TXT/Map).
"""

import io
import json
import math
import os
import time
from datetime import datetime

from flask import Blueprint, abort, make_response, render_template, request, send_from_directory, jsonify

from modules.state import STATE
from modules.utils import sanitize_point_name, robust_avg, robust_avg_stats
from modules.survey import (
    list_survey_ids, load_survey, save_survey, create_survey,
    delete_survey_file, next_point_id, point_feature,
    flatten_point_for_csv, CSV_HEADER, SURVEY_DIR, survey_path, point_from_feature,
    backup_survey, survey_audio_dir, build_voice_notes_index, move_pending_notes_to_feature,
    remove_note_by_id, find_voice_note, cleanup_note_file, get_survey_dir
)
from modules.exports import (
    build_dxf_advanced, export_geopackage, format_point_txt
)
from modules.active_survey import get_active_survey_id, set_active_survey_id
from modules.codici_punto import load_codici
from modules import settings as _settings
from modules.session_log import log_event, read_log
from modules.geodesy import WGS84_A

bp = Blueprint('surveys', __name__)

# GNSS data freshness thresholds (seconds)
_GNSS_STALE_WARN_S = 5.0   # age below this → run quality gate checks
_GNSS_STALE_ERROR_S = 10.0  # age above this → reject measurement entirely


def _redirect(url: str):
    r = make_response("", 302)
    r.headers["Location"] = url
    return r


def _tpv_age_seconds(snap: dict):
    """Return age of TPV data in seconds, or None if no timestamp."""
    tpv = snap.get("TPV", {})
    ts_str = tpv.get("time")
    if not ts_str:
        return None
    try:
        ts = datetime.fromisoformat(ts_str)
        return (datetime.now() - ts).total_seconds()
    except Exception:
        return None


def _json_ok(**payload):
    payload.setdefault("ok", True)
    return make_response(json.dumps(payload, ensure_ascii=False), 200, {"Content-Type": "application/json; charset=utf-8"})


def _json_err(message: str, status: int = 400, **payload):
    payload.setdefault("ok", False)
    payload.setdefault("error", message)
    return make_response(json.dumps(payload, ensure_ascii=False), status, {"Content-Type": "application/json; charset=utf-8"})


def _current_gnss_snapshot():
    snap = STATE.snapshot()
    tpv = snap.get("TPV", {})
    hp = snap.get("HPPOSLLH", {})
    dop = snap.get("DOP", {})
    return {
        "time": tpv.get("time"),
        "mode": tpv.get("mode"),
        "rtk": tpv.get("rtk"),
        "numSV": tpv.get("numSV"),
        "lat": hp.get("lat") if hp.get("lat") is not None else tpv.get("lat"),
        "lon": hp.get("lon") if hp.get("lon") is not None else tpv.get("lon"),
        "altHAE": hp.get("altHAE"),
        "altMSL": hp.get("altMSL") if hp.get("altMSL") is not None else tpv.get("altMSL"),
        "hAcc": hp.get("hAcc") if hp.get("hAcc") is not None else tpv.get("hAcc"),
        "vAcc": hp.get("vAcc") if hp.get("vAcc") is not None else tpv.get("vAcc"),
        "pdop": dop.get("pdop"),
        "hdop": dop.get("hdop"),
        "vdop": dop.get("vdop"),
    }


def _new_note_payload(sid: str, kind: str, filename: str, point_name: str = "", point_code: str = "", client_note: str = ""):
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return {
        "id": f"vn_{stamp}_{int(time.time()*1000)%1000:03d}",
        "kind": kind,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "audio_filename": filename,
        "audio_file": f"media/{sid}/audio/{filename}",
        "audio_url": f"/survey/{sid}/media/audio/{filename}",
        "point_id": None,
        "point_name": (point_name or "").strip()[:40],
        "point_code": (point_code or "").strip()[:20],
        "note": (client_note or "").strip()[:300],
        "transcript": "",
        "transcript_status": "pending",
        "gnss": _current_gnss_snapshot(),
    }



# ---------- List ----------
@bp.route("/surveys")
def surveys_list():
    ids = list_survey_ids()
    active_sid = get_active_survey_id()
    active_title = None
    items = []
    if not ids:
        items.append('<div class="card">Nessun rilievo. <a class="btn" href="/survey/new">Crea il primo</a></div>')
    for sid in ids:
        try:
            svy = load_survey(sid)
            title = svy.get("properties", {}).get("title", sid)
            desc = svy.get("properties", {}).get("desc", "")
            npt = len(svy.get("features", []))
            if sid == active_sid:
                active_title = title
            is_active = (sid == active_sid)
            set_btn = (
                f'<button class="btn btn-set-active" data-sid="{sid}">📌 Imposta attivo</button>'
                if not is_active else
                '<span class="active-badge">✅ Attivo</span>'
            )
            items.append(
                f'<div class="card"><div class="kv"><span><b>{title}</b> ({sid})</span><span>{npt} punti</span></div>'
                f'<div class="kv"><span>{desc or "-"}</span>'
                f'<span><a class="btn" href="/survey/{sid}">Apri</a> {set_btn}</span></div></div>'
            )
        except Exception as e:
            items.append(f'<div class="card">Errore con {sid}: {e}</div>')
    return render_template('rtk_surveys_list.html',
                           items="\n".join(items),
                           active_sid=active_sid,
                           active_title=active_title)




# ---------- New survey ----------
@bp.route("/survey/new", methods=["GET", "POST"])
def survey_new():
    if request.method == "GET":
        return render_template('rtk_survey_new.html')
    title = (request.form.get("title") or "").strip()[:40]
    desc = (request.form.get("desc") or "").strip()[:1000]
    sid = create_survey(title, desc)
    set_active_survey_id(sid)
    return _redirect(f"/survey/{sid}")


# ---------- Active survey ----------
@bp.route("/api/survey/<sid>/set_active", methods=["POST"])
def survey_set_active(sid):
    try:
        load_survey(sid)
    except FileNotFoundError:
        return make_response(json.dumps({"ok": False, "error": "not found"}), 404,
                             {"Content-Type": "application/json"})
    set_active_survey_id(sid)
    return make_response(json.dumps({"ok": True, "sid": sid}), 200,
                         {"Content-Type": "application/json"})


@bp.route("/survey/active/point")
def survey_active_point():
    active_sid = get_active_survey_id()
    if not active_sid:
        return _redirect("/surveys")
    return _redirect(f"/survey/{active_sid}/point")



@bp.route("/survey/<sid>")
def survey_view(sid):
    try:
        svy = load_survey(sid)
    except FileNotFoundError:
        abort(404)
    props = svy.get("properties", {})
    active_sid = get_active_survey_id()
    rows = []
    points_for_json = []
    point_cards = []
    for idx, f in enumerate(svy.get("features", [])):
        p = f.get("properties", {})

        def fnum(x, fmt):
            return "-" if x is None else fmt.format(x)

        point_id = f.get('id', '')
        point_name = p.get('name', '')
        codice = p.get('codice', '')
        points_for_json.append({"id": point_id, "name": point_name})

        sigma_N = p.get("sigma_n")
        sigma_E = p.get("sigma_e")
        sigma_U = p.get("sigma_u")

        def fsig(x):
            return f"±{x:.3f} m" if x is not None else "-"

        rows.append(
            "<tr>"
            f"<td><input type='checkbox' class='area-cb' value='{point_id}'/></td>"
            f"<td>{point_id}</td>"
            f"<td>{point_name}</td>"
            f"<td>{codice}</td>"
            f"<td>{fnum(p.get('lat'), '{:.9f}')}</td>"
            f"<td>{fnum(p.get('lon'), '{:.9f}')}</td>"
            f"<td>{fnum(p.get('alt_hae'), '{:.3f}')}</td>"
            f"<td>{fnum(p.get('pdop'), '{:.2f}')}</td>"
            f"<td>{fnum(p.get('hdop'), '{:.2f}')}</td>"
            f"<td>{fnum(p.get('vdop'), '{:.2f}')}</td>"
            f"<td>{fnum(p.get('rel_n'), '{:.4f}')}</td>"
            f"<td>{fnum(p.get('rel_e'), '{:.4f}')}</td>"
            f"<td>{fnum(p.get('rel_d'), '{:.4f}')}</td>"
            f"<td>{fnum(p.get('baseline'), '{:.4f}')}</td>"
            f"<td>{fsig(sigma_N)}</td>"
            f"<td>{fsig(sigma_E)}</td>"
            f"<td>{fsig(sigma_U)}</td>"
            f"<td><a class='btn' href='/survey/{sid}/point/{point_id}.txt'>TXT</a></td>"
            f"<td><button class='btn-delete-point' onclick=\"confirmDeletePoint('{sid}', {idx}, '{point_name}')\">🗑️</button></td>"
            "</tr>"
        )
        point_cards.append({
            "idx": idx,
            "id": point_id,
            "name": point_name,
            "codice": codice,
            "lat": fnum(p.get('lat'), '{:.9f}'),
            "lon": fnum(p.get('lon'), '{:.9f}'),
            "hae": fnum(p.get('alt_hae'), '{:.3f}'),
            "pdop": fnum(p.get('pdop'), '{:.2f}'),
            "hdop": fnum(p.get('hdop'), '{:.2f}'),
            "vdop": fnum(p.get('vdop'), '{:.2f}'),
            "sigma_n": fsig(sigma_N),
            "sigma_e": fsig(sigma_E),
            "sigma_u": fsig(sigma_U),
        })
    num_points = len(rows)
    num_columns = 19
    voice_notes = build_voice_notes_index(sid, svy)
    return render_template('rtk_survey_view.html',
                           sid=props.get("id", sid),
                           title=props.get("title", sid),
                           notes=props.get("desc", ""),
                           rows="\n".join(rows) or f"<tr><td colspan='{num_columns}'>(nessun punto)</td></tr>",
                           points_json=json.dumps(points_for_json),
                           num_points=str(num_points),
                           active_sid=active_sid,
                           voice_notes=voice_notes,
                           point_cards=point_cards,
                           gnss_now=_current_gnss_snapshot())


@bp.route("/survey/<sid>/media/audio/<path:filename>")
def survey_media_audio(sid, filename):
    directory = survey_audio_dir(sid)
    return send_from_directory(directory, filename, as_attachment=False)


@bp.route("/survey/<sid>/voice-note", methods=["POST"])
def survey_voice_note_create(sid):
    try:
        svy = load_survey(sid)
    except FileNotFoundError:
        abort(404)
    file = request.files.get("audio")
    if not file or not file.filename:
        return _json_err("Nessun file audio ricevuto")
    kind = (request.form.get("kind") or "point").strip().lower()
    if kind not in ("point", "session"):
        kind = "point"
    ext = os.path.splitext(file.filename)[1].lower() or ".webm"
    if ext not in (".webm", ".wav", ".ogg", ".mp3", ".m4a"):
        ext = ".webm"
    filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{int(time.time()*1000)%1000:03d}_{kind}{ext}"
    path = os.path.join(survey_audio_dir(sid), filename)
    file.save(path)
    note = _new_note_payload(
        sid=sid,
        kind=kind,
        filename=filename,
        point_name=request.form.get("point_name", ""),
        point_code=request.form.get("point_code", ""),
        client_note=request.form.get("note", ""),
    )
    props = svy.setdefault("properties", {})
    if kind == "session":
        props.setdefault("voice_notes_session", []).append(note)
    else:
        props.setdefault("voice_notes_pending", []).append(note)
    backup_survey(sid)
    save_survey(sid, svy)
    return _json_ok(note=note)


@bp.route("/survey/<sid>/voice-note/<note_id>", methods=["POST"])
def survey_voice_note_update(sid, note_id):
    try:
        svy = load_survey(sid)
    except FileNotFoundError:
        abort(404)
    kind, notes, idx, note, feat = find_voice_note(svy, note_id)
    if note is None:
        return _json_err("Nota non trovata", 404)
    note["note"] = (request.form.get("note") or note.get("note", "")).strip()[:300]
    if kind in ("voice_notes_session", "voice_notes_pending") or kind == "feature":
        if request.form.get("point_name") is not None:
            note["point_name"] = (request.form.get("point_name") or "").strip()[:40]
        if request.form.get("point_code") is not None:
            note["point_code"] = (request.form.get("point_code") or "").strip()[:20]
    backup_survey(sid)
    save_survey(sid, svy)
    return _json_ok(note=note)


@bp.route("/survey/<sid>/voice-note/<note_id>/delete", methods=["POST"])
def survey_voice_note_delete(sid, note_id):
    try:
        svy = load_survey(sid)
    except FileNotFoundError:
        abort(404)
    if not remove_note_by_id(sid, svy, note_id):
        return _json_err("Nota non trovata", 404)
    backup_survey(sid)
    save_survey(sid, svy)
    return _json_ok(message="Nota eliminata")


# ---------- Update notes ----------
@bp.route("/survey/<sid>/notes", methods=["POST"])
def survey_update_notes(sid):
    try:
        svy = load_survey(sid)
    except FileNotFoundError:
        abort(404)
    notes = (request.form.get("notes") or "").strip()[:1000]
    svy.setdefault("properties", {})["desc"] = notes
    save_survey(sid, svy)
    return _redirect(f"/survey/{sid}")


# ---------- Add point ----------
@bp.route("/survey/<sid>/point", methods=["GET", "POST"])
def survey_point(sid):
    if request.method == "GET":
        try:
            svy = load_survey(sid)
        except FileNotFoundError:
            abort(404)
        next_pid = next_point_id(svy)
        codici_data = load_codici()
        active_title = svy.get("properties", {}).get("title", sid)
        num_points = len(svy.get("features", []))
        saved = request.args.get("saved", "")
        savedname = request.args.get("savedname", "")
        savedcodice = request.args.get("savedcodice", "")
        return render_template('rtk_survey_point_form.html',
                               sid=sid,
                               next_pid=next_pid,
                               active_title=active_title,
                               num_points=num_points,
                               codici_json=json.dumps(codici_data),
                               saved=saved,
                               savedname=savedname,
                               savedcodice=savedcodice,
                               gnss_now=_current_gnss_snapshot())

    name = sanitize_point_name(request.form.get("name", ""))
    desc = (request.form.get("desc", "") or "").strip()[:300]
    codice = (request.form.get("codice", "") or "").strip()[:20]
    dur_form = request.form.get("dur", "").strip()
    force = request.form.get("force", "0") == "1"
    duration = None
    if dur_form:
        try:
            duration = float(dur_form)
        except Exception:
            duration = None
    if duration is None:
        try:
            duration = float(request.args.get("dur", 10.0))
        except Exception:
            duration = 10.0
    try:
        interval = float(request.args.get("step", 0.5))
    except Exception:
        interval = 0.5

    # ---------- Pre-sampling GNSS checks ----------
    if not force:
        pre_snap = STATE.snapshot()
        age = _tpv_age_seconds(pre_snap)

        # Check for stale or missing GNSS data (>_GNSS_STALE_ERROR_S seconds)
        if age is None or age > _GNSS_STALE_ERROR_S:
            try:
                log_event("point_failed", sid, {"reason": "no_gnss_data"})
            except Exception:
                pass
            return make_response(json.dumps({
                "ok": False,
                "error": "no_gnss_data",
                "message": "Nessun dato GNSS ricevuto. Verificare la connessione."
            }), 200, {"Content-Type": "application/json"})

        # Quality gate checks (only when GNSS data is fresh, i.e., age < _GNSS_STALE_WARN_S)
        if age < _GNSS_STALE_WARN_S:
            cfg = _settings.load_settings()
            if cfg.get("rtk_quality_gate", True):
                tpv_pre = pre_snap.get("TPV", {})
                hp_pre = pre_snap.get("HPPOSLLH", {})
                dop_pre = pre_snap.get("DOP", {})
                warnings = []

                rtk_val = tpv_pre.get("rtk", "none")
                if rtk_val == "none":
                    warnings.append(f"Nessun fix RTK (stato: {rtk_val})")

                hacc = hp_pre.get("hAcc")
                max_hacc = cfg.get("max_hacc", 0.05)
                if hacc is not None and hacc > max_hacc:
                    warnings.append(
                        f"Precisione orizzontale elevata: {hacc:.3f} m (soglia: {max_hacc} m)"
                    )

                pdop = dop_pre.get("pdop")
                max_pdop = cfg.get("max_pdop", 3.0)
                if pdop is not None and pdop > max_pdop:
                    warnings.append(f"PDOP elevato: {pdop:.2f} (soglia: {max_pdop})")

                numsv = tpv_pre.get("numSV")
                min_sv = cfg.get("min_sv", 8)
                if numsv is not None and numsv < min_sv:
                    warnings.append(f"Pochi satelliti: {numsv} (minimo: {min_sv})")

                if warnings:
                    return make_response(json.dumps({
                        "ok": False,
                        "quality_warning": True,
                        "warnings": warnings
                    }), 200, {"Content-Type": "application/json"})

    start_ts = datetime.now()
    samples = []
    n_iters = max(1, int(duration / interval))
    for _ in range(n_iters):
        samples.append(STATE.snapshot())
        time.sleep(interval)
    end_ts = datetime.now()

    def collect_key(group, key):
        vals = []
        for s in samples:
            g = s.get(group, {})
            v = g.get(key, None)
            if isinstance(v, (int, float)):
                vals.append(v)
        return robust_avg(vals)

    mode = int(round(collect_key("TPV", "mode") or 0))
    rtk = samples[-1].get("TPV", {}).get("rtk", "-")
    numSV = int(round(collect_key("TPV", "numSV") or 0))

    lat = collect_key("HPPOSLLH", "lat") or collect_key("TPV", "lat")
    lon = collect_key("HPPOSLLH", "lon") or collect_key("TPV", "lon")
    altHAE = collect_key("HPPOSLLH", "altHAE")
    altMSL = collect_key("HPPOSLLH", "altMSL") or collect_key("TPV", "altMSL")
    hAcc = collect_key("HPPOSLLH", "hAcc") or collect_key("TPV", "hAcc")
    vAcc = collect_key("HPPOSLLH", "vAcc") or collect_key("TPV", "vAcc")

    X = collect_key("HPPOSECEF", "X")
    Y = collect_key("HPPOSECEF", "Y")
    Z = collect_key("HPPOSECEF", "Z")
    pAcc = collect_key("HPPOSECEF", "pAcc")

    gdop = collect_key("DOP", "gdop"); pdop = collect_key("DOP", "pdop")
    hdop = collect_key("DOP", "hdop"); vdop = collect_key("DOP", "vdop")
    ndop = collect_key("DOP", "ndop"); edop = collect_key("DOP", "edop")
    tdop = collect_key("DOP", "tdop")

    covNN = collect_key("COV", "covNN"); covEE = collect_key("COV", "covEE")
    covDD = collect_key("COV", "covDD"); covNE = collect_key("COV", "covNE")
    covND = collect_key("COV", "covND"); covED = collect_key("COV", "covED")

    relN = collect_key("RELPOS", "N"); relE = collect_key("RELPOS", "E")
    relD = collect_key("RELPOS", "D")
    relsN = collect_key("RELPOS", "sN"); relsE = collect_key("RELPOS", "sE")
    relsD = collect_key("RELPOS", "sD")
    horiz = baseline = bearing = slope = None
    if relN is not None and relE is not None and relD is not None:
        horiz = math.hypot(relN, relE)
        baseline = math.sqrt(horiz * horiz + relD * relD)
        bearing = math.degrees(math.atan2(relE, relN))
        bearing = bearing + 360 if bearing < 0 else bearing
        slope = math.degrees(math.atan2(-relD, horiz)) if horiz else 0.0

    # ---------- Sigma (std dev) of position samples ----------
    def collect_vals(group, key):
        vals = []
        for s in samples:
            g = s.get(group, {})
            v = g.get(key, None)
            if isinstance(v, (int, float)):
                vals.append(v)
        return vals

    _lat_stats = robust_avg_stats(collect_vals("HPPOSLLH", "lat"))
    _lon_stats = robust_avg_stats(collect_vals("HPPOSLLH", "lon"))
    _alt_stats = robust_avg_stats(collect_vals("HPPOSLLH", "altHAE"))

    _mean_lat_rad = math.radians(lat or 0.0)
    sigma_N = None
    sigma_E = None
    sigma_U = None
    n_kept = _alt_stats.get("n_kept", 0)
    if _lat_stats.get("sigma") is not None:
        sigma_N = _lat_stats["sigma"] * math.radians(1.0) * WGS84_A
    if _lon_stats.get("sigma") is not None:
        sigma_E = _lon_stats["sigma"] * math.radians(1.0) * WGS84_A * math.cos(_mean_lat_rad)
    if _alt_stats.get("sigma") is not None:
        sigma_U = _alt_stats["sigma"]

    stats = {
        "mode": mode, "rtk": rtk, "numSV": numSV,
        "hAcc": hAcc, "vAcc": vAcc, "pAcc": pAcc,
        "gdop": gdop, "pdop": pdop, "hdop": hdop, "vdop": vdop,
        "ndop": ndop, "edop": edop, "tdop": tdop,
        "covNN": covNN, "covEE": covEE, "covDD": covDD,
        "covNE": covNE, "covND": covND, "covED": covED,
        "relN": relN, "relE": relE, "relD": relD,
        "relsN": relsN, "relsE": relsE, "relsD": relsD,
        "baseline": baseline, "horiz": horiz, "bearing": bearing, "slope": slope,
        "sigma_N": sigma_N, "sigma_E": sigma_E, "sigma_U": sigma_U, "n_kept": n_kept,
    }

    try:
        svy = load_survey(sid)
    except FileNotFoundError:
        abort(404)
    pid = next_point_id(svy)
    start_iso = start_ts.isoformat(timespec='seconds')
    end_iso = end_ts.isoformat(timespec='seconds')

    feat = point_feature(
        pid, lat, lon, altHAE, altMSL, X, Y, Z, stats,
        {"name": name, "codice": codice, "desc": desc, "duration": duration,
         "interval": interval, "n_samples": len(samples),
         "start": start_iso, "end": end_iso}
    )
    move_pending_notes_to_feature(svy, feat, name, codice)
    svy.setdefault("features", []).append(feat)
    backup_survey(sid)
    save_survey(sid, svy)

    try:
        log_event("point_saved", sid, {"pid": pid, "codice": codice, "rtk": rtk})
    except Exception:
        pass

    return _redirect(f"/survey/{sid}/point?saved={pid}&savedname={name}&savedcodice={codice}")


# ---------- Downloads ----------
@bp.route("/survey/<sid>.geojson")
def survey_download_geojson(sid):
    path = survey_path(sid)
    if not os.path.isfile(path):
        abort(404)
    return send_from_directory(get_survey_dir(), os.path.basename(path),
                               as_attachment=True, download_name=f"{sid}.geojson")


@bp.route("/survey/<sid>.xlsx")
def survey_download_xlsx(sid):
    try:
        svy = load_survey(sid)
    except FileNotFoundError:
        abort(404)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return make_response("openpyxl non installato. Eseguire: pip install openpyxl", 500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sid[:31]  # Excel limita il nome foglio a 31 caratteri

    # Intestazione
    headers = ["id"] + CSV_HEADER
    header_fill = PatternFill(start_color="1D5F8F", end_color="1D5F8F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Formati numerici
    FMT_9 = '0.000000000'   # 9 decimali per lat/lon
    FMT_4 = '0.0000'        # 4 decimali
    FMT_3 = '0.000'         # 3 decimali per quote
    FMT_2 = '0.00'          # 2 decimali
    FMT_1 = '0.0'           # 1 decimale
    FMT_SCI = '0.0000E+00'  # scientifico per covarianze
    FMT_INT = '0'           # intero

    # Mappa colonna -> formato (basato su CSV_HEADER, offset +2 perché col 1 è "id")
    col_formats = {
        "lat": FMT_9, "lon": FMT_9,
        "alt_hae": FMT_3, "alt_msl": FMT_3,
        "ecef_x": FMT_4, "ecef_y": FMT_4, "ecef_z": FMT_4,
        "gdop": FMT_2, "pdop": FMT_2, "hdop": FMT_2, "vdop": FMT_2,
        "ndop": FMT_2, "edop": FMT_2, "tdop": FMT_2,
        "h_acc": FMT_3, "v_acc": FMT_3, "p_acc": FMT_3,
        "cov_nn": FMT_SCI, "cov_ee": FMT_SCI, "cov_dd": FMT_SCI,
        "cov_ne": FMT_SCI, "cov_nd": FMT_SCI, "cov_ed": FMT_SCI,
        "rel_n": FMT_4, "rel_e": FMT_4, "rel_d": FMT_4,
        "rel_sn": FMT_3, "rel_se": FMT_3, "rel_sd": FMT_3,
        "baseline": FMT_4, "horiz": FMT_4,
        "bearing_deg": FMT_2, "slope_deg": FMT_2,
        "sigma_n": FMT_4, "sigma_e": FMT_4, "sigma_u": FMT_4,
        "duration_s": FMT_1, "interval_s": FMT_2,
        "gnss_mode": FMT_INT, "num_sv": FMT_INT,
        "n_kept": FMT_INT, "n_samples": FMT_INT,
    }

    # Dati
    for row_idx, feat in enumerate(svy.get("features", []), 2):
        p = feat.get("properties", {})
        pid = feat.get("id", "")

        def fv(key):
            """Restituisce il valore nativo (float/int/str/None) per la cella."""
            v = p.get(key)
            if v is None:
                return None
            # Campi stringa
            if key in ("name", "codice", "desc", "timestamp", "rtk", "start_time", "end_time"):
                return str(v)
            # Campi numerici interi
            if key in ("gnss_mode", "num_sv", "n_kept", "n_samples"):
                try:
                    return int(v)
                except (TypeError, ValueError):
                    return None
            # Tutti gli altri: float
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        # Colonna 1: id
        ws.cell(row=row_idx, column=1, value=pid)

        # Colonne dati
        for col_idx, key in enumerate(CSV_HEADER, 2):
            val = fv(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            # Applicare formato numerico
            fmt = col_formats.get(key)
            if fmt and val is not None and isinstance(val, (int, float)):
                cell.number_format = fmt

    # Larghezza colonne automatica (approssimativa)
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(10, len(header) + 2)

    # Freeze prima riga
    ws.freeze_panes = "A2"

    # Salva in buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = make_response(buf.read())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f'attachment; filename="{sid}.xlsx"'
    return resp


@bp.route("/survey/<sid>/map")
def survey_map(sid):
    try:
        svy = load_survey(sid)
    except FileNotFoundError:
        abort(404)
    if not svy.get("features"):
        return make_response("Nessun punto da visualizzare", 400)
    return render_template('rtk_map.html', sid=sid, geojson=json.dumps(svy))


@bp.route("/survey/<sid>/cad")
def survey_cad(sid):
    """Open the CAD editor pre-loaded with this survey."""
    try:
        svy = load_survey(sid)
    except FileNotFoundError:
        abort(404)
    title = svy.get("properties", {}).get("title", sid)
    return render_template('rtk_survey_cad.html', sid=sid, title=title)


@bp.route("/survey/<sid>.dxf")
def survey_download_dxf(sid):
    try:
        svy = load_survey(sid)
    except FileNotFoundError:
        abort(404)
    try:
        text_height = float(request.args.get("text_height", "0.1"))
        show_precision = request.args.get("precision", "false").lower() in ("true", "1", "yes")
        quota_decimals = int(request.args.get("decimals", "3"))
        dxf_text = build_dxf_advanced(svy, mode="3d", text_height=text_height,
                                      show_precision=show_precision, quota_decimals=quota_decimals)
    except Exception as e:
        return make_response(f"Errore DXF: {e}", 500)
    resp = make_response(dxf_text)
    resp.headers["Content-Type"] = "application/dxf; charset=utf-8"
    resp.headers["Content-Disposition"] = f'attachment; filename="{sid}.dxf"'
    return resp


@bp.route("/survey/<sid>_2d.dxf")
def survey_download_dxf_2d(sid):
    try:
        svy = load_survey(sid)
    except FileNotFoundError:
        abort(404)
    try:
        text_height = float(request.args.get("text_height", "0.1"))
        show_precision = request.args.get("precision", "false").lower() in ("true", "1", "yes")
        quota_decimals = int(request.args.get("decimals", "3"))
        dxf_text = build_dxf_advanced(svy, mode="2d", text_height=text_height,
                                      show_precision=show_precision, quota_decimals=quota_decimals)
    except Exception as e:
        return make_response(f"Errore DXF: {e}", 500)
    resp = make_response(dxf_text)
    resp.headers["Content-Type"] = "application/dxf; charset=utf-8"
    resp.headers["Content-Disposition"] = f'attachment; filename="{sid}_2d.dxf"'
    return resp


@bp.route("/survey/<sid>.gpkg")
def survey_download_geopackage(sid):
    try:
        svy = load_survey(sid)
    except FileNotFoundError:
        abort(404)
    if not svy.get("features"):
        return make_response("Nessun punto nel rilievo", 400)
    filepath = export_geopackage(svy, sid)
    if not filepath:
        return make_response("Errore nella creazione del GeoPackage", 500)
    try:
        with open(filepath, "rb") as fh:
            gpkg_data = fh.read()
        try:
            os.unlink(filepath)
        except Exception:
            pass
        resp = make_response(gpkg_data)
        resp.headers["Content-Type"] = "application/geopackage+sqlite3"
        resp.headers["Content-Disposition"] = f'attachment; filename="{sid}.gpkg"'
        return resp
    except Exception as e:
        return make_response(f"Errore lettura GeoPackage: {e}", 500)


@bp.route("/survey/<sid>/point/<pid>.txt")
def survey_point_txt(sid, pid):
    try:
        svy = load_survey(sid)
    except FileNotFoundError:
        abort(404)
    feat = next((f for f in svy.get("features", []) if f.get("id") == pid), None)
    if not feat:
        abort(404)
    txt = format_point_txt(feat, sid)
    resp = make_response(txt)
    resp.headers["Content-Type"] = "text/plain; charset=utf-8"
    resp.headers["Content-Disposition"] = f'attachment; filename="{sid}-{pid}.txt"'
    return resp


# ---------- Delete ----------
@bp.route("/survey/<sid>/point/<int:point_index>/delete", methods=["POST"])
def delete_point(sid, point_index):
    try:
        svy = load_survey(sid)
    except FileNotFoundError:
        return make_response({"error": "Survey not found"}, 404)
    features = svy.get("features", [])
    if point_index < 0 or point_index >= len(features):
        return make_response({"error": f"Point not found at index {point_index}"}, 404)
    removed = features.pop(point_index)
    for note in list(removed.get("properties", {}).get("voice_notes", []) or []):
        cleanup_note_file(sid, note)
    save_survey(sid, svy)
    return jsonify({
        "success": True,
        "message": f"Punto {removed.get('id', '')} eliminato"
    }), 200



@bp.route("/survey/<sid>/delete", methods=["POST"])
def delete_survey(sid):
    try:
        load_survey(sid)
    except FileNotFoundError:
        return jsonify({"success": False, "error": "Survey not found"}), 404

    data = request.get_json(silent=True) or {}
    confirm = (data.get("confirmation") or "").strip()

    if confirm != "ELIMINA":
        return jsonify({
            "success": False,
            "error": "Conferma non valida. Scrivi 'ELIMINA' per confermare."
        }), 400

    delete_survey_file(sid)
    return jsonify({
        "success": True,
        "message": f"Rilievo {sid} eliminato"
    }), 200


# ---------- Area ----------
@bp.route("/survey/<sid>/area", methods=["POST"])
def survey_calc_area(sid):
    from modules.cogo import calc_area_perimeter
    from modules.geodesy import geodetic_to_ecef, ecef_delta_to_enu

    data = request.get_json()
    if not data:
        return make_response({"error": "No data"}, 400)
    point_ids = data.get("point_ids") or data.get("points", [])
    if len(point_ids) < 3:
        return make_response({"error": "Servono almeno 3 punti"}, 400)

    try:
        svy = load_survey(sid)
    except FileNotFoundError:
        return make_response({"error": "Survey not found"}, 404)

    features = svy.get("features", [])
    ordered_feats = []
    for pid in point_ids:
        f = next((f for f in features if f.get("id") == pid), None)
        if f:
            ordered_feats.append(f)

    if len(ordered_feats) < 3:
        return make_response({"error": "Punti insufficienti con coordinate valide"}, 400)

    # Get reference point
    first_p = ordered_feats[0].get("properties", {})
    lat0, lon0 = first_p.get("lat"), first_p.get("lon")
    if lat0 is None or lon0 is None:
        return make_response({"error": "Coordinate primo punto non valide"}, 400)
    alt0 = first_p.get("alt_msl") or first_p.get("alt_hae", 0.0)
    X0, Y0, Z0 = geodetic_to_ecef(lat0, lon0, alt0)

    enu_points = []
    for f in ordered_feats:
        p = f.get("properties", {})
        lat, lon = p.get("lat"), p.get("lon")
        if lat is None or lon is None:
            continue
        alt = p.get("alt_msl") or p.get("alt_hae", 0.0)
        X, Y, Z = geodetic_to_ecef(lat, lon, alt)
        e, n, u = ecef_delta_to_enu(X - X0, Y - Y0, Z - Z0, lat0, lon0)
        enu_points.append((e, n))

    if len(enu_points) < 3:
        return make_response({"error": "Punti insufficienti con coordinate valide"}, 400)

    result = calc_area_perimeter(enu_points)
    return make_response(result, 200)


# ---------- Compat / old routes ----------
@bp.route("/survey", methods=["GET"])
def survey_redirect():
    return _redirect("/surveys")


# ---------- Session log ----------
@bp.route("/api/session_log")
def api_session_log():
    try:
        limit = int(request.args.get("limit", 200))
    except (ValueError, TypeError):
        limit = 200
    events = read_log(limit=limit)
    return make_response(json.dumps(events, ensure_ascii=False),
                         200, {"Content-Type": "application/json"})
